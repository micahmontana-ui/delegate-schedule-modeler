"""
Module 7a: Excel Workbook Builder
"""

from __future__ import annotations

import math
from collections import defaultdict
from datetime import date
from typing import Any

import pandas as pd


# ─── Brand colours ────────────────────────────────────────────────────────────
_DARK_TEAL   = "015480"   # dark header bg
_MED_BLUE    = "0F91BB"   # accent header bg
_LIGHT_BLUE  = "6EC5EA"   # section sub-header
_RUST        = "C65B41"   # warning / negative
_AMBER       = "FFB31C"   # highlight
_SAGE        = "69B6AF"   # positive / good
_WHITE       = "FFFFFF"
_NEAR_WHITE  = "EAF6FB"   # alternating row tint
_LIGHT_GREY  = "D9D9D9"


def build_workbook(
    output_path: str,
    df: pd.DataFrame,
    bus_trips_df: pd.DataFrame,
    stop_summary: dict[str, dict],
    non_attendance: pd.DataFrame,
    required_activities: dict[str, dict],
    optional_activities: dict[str, dict],
    hotels: list[dict[str, Any]],
    assumptions: list[str],
    smpw_df: pd.DataFrame | None = None,
) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
    except ImportError:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Style helpers ──────────────────────────────────────────────────────────
    def _fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color=_WHITE, size=11, italic=False):
        return Font(bold=bold, color=color, size=size, italic=italic)

    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    def _thin_border():
        s = Side(style="thin", color="C0C0C0")
        return Border(left=s, right=s, top=s, bottom=s)

    def _style_header_row(ws, row_num=1, fill_hex=_DARK_TEAL):
        for cell in ws[row_num]:
            cell.fill = _fill(fill_hex)
            cell.font = _font(bold=True, color=_WHITE)
            cell.alignment = _align(h="center")

    def _auto_col_widths(ws, min_w=8, max_w=45):
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            best = min_w
            for cell in col:
                try:
                    best = max(best, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(best + 2, max_w)

    def _freeze(ws, cell="A2"):
        ws.freeze_panes = cell

    def _write_sheet(ws, headers, rows, fill_hex=_DARK_TEAL, alt_rows=True):
        ws.append(headers)
        _style_header_row(ws, 1, fill_hex)
        for i, row in enumerate(rows):
            ws.append(row)
            if alt_rows and i % 2 == 0:
                for cell in ws[ws.max_row]:
                    cell.fill = _fill(_NEAR_WHITE)
                    cell.font = _font(color="000000")
            else:
                for cell in ws[ws.max_row]:
                    cell.font = _font(color="000000")
        _auto_col_widths(ws)
        _freeze(ws)

    # ── Derived data helpers ───────────────────────────────────────────────────
    req_names  = list(required_activities.keys())
    opt_names  = list(optional_activities.keys())
    act_cols   = [f"req_{n}" for n in req_names] + [f"opt_{n}" for n in opt_names]

    def _col_val(row, col):
        if col not in df.columns:
            return None
        v = row.get(col)
        return None if pd.isna(v) else v

    # Collect all unique activity dates
    all_dates_set: set[date] = set()
    for col in act_cols:
        if col in df.columns:
            for v in df[col].dropna():
                d = v if isinstance(v, date) else pd.Timestamp(v).date()
                all_dates_set.add(d)
    if smpw_df is not None and not smpw_df.empty:
        for v in smpw_df["SMPW_Date"].dropna():
            d = v if isinstance(v, date) else pd.Timestamp(v).date()
            all_dates_set.add(d)
    all_dates = sorted(all_dates_set)

    # ── Sheet: Summary ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.sheet_properties.tabColor = _DARK_TEAL

    def _sum_row(label, value, bold=False):
        ws_sum.append([label, value])
        r = ws_sum.max_row
        ws_sum.cell(r, 1).font = _font(bold=bold, color="000000")
        ws_sum.cell(r, 2).font = _font(bold=bold, color="000000")
        ws_sum.cell(r, 2).alignment = _align(h="center")

    def _sum_section(title):
        ws_sum.append([title])
        r = ws_sum.max_row
        ws_sum.cell(r, 1).fill = _fill(_DARK_TEAL)
        ws_sum.cell(r, 1).font = _font(bold=True, color=_WHITE, size=12)
        ws_sum.merge_cells(f"A{r}:B{r}")

    # Header
    ws_sum["A1"] = "Convention Logistics — Results Summary"
    ws_sum["A1"].font = _font(bold=True, color=_WHITE, size=14)
    ws_sum["A1"].fill = _fill(_MED_BLUE)
    ws_sum.merge_cells("A1:B1")
    ws_sum["A1"].alignment = _align(h="center")
    ws_sum.append([])

    _sum_section("Population")
    _sum_row("Total Groups", len(df))
    _sum_row("Total Delegates", int(df["GroupSize"].sum()))
    ws_sum.append([])

    _sum_section("Activity Assignment")
    for name in req_names:
        col = f"req_{name}"
        n = int(df[col].notna().sum()) if col in df.columns else 0
        rate = f"{100*n//len(df)}%" if len(df) else "—"
        _sum_row(f"{name} — Assigned", f"{n:,}  ({rate})")
        _sum_row(f"{name} — Not Scheduled", f"{len(df)-n:,}")
    if smpw_df is not None and not smpw_df.empty:
        _sum_row("SMPW Groups", len(smpw_df))
        _sum_row("SMPW Delegates", int(smpw_df["GroupSize"].sum()))
    ws_sum.append([])

    _sum_section("Bus Performance")
    total_trips = len(bus_trips_df) if not bus_trips_df.empty else 0
    under_min = int(bus_trips_df["UnderMinimum"].sum()) if not bus_trips_df.empty else 0
    pct_under = f"{100*under_min/total_trips:.1f}%" if total_trips else "0%"
    _sum_row("Total Bus Trips", total_trips)
    _sum_row("Under-Minimum Trips", f"{under_min}  ({pct_under})")

    ws_sum.column_dimensions["A"].width = 38
    ws_sum.column_dimensions["B"].width = 20

    # ── Sheet: Daily Activity Summary (overall pivot) ──────────────────────────
    ws_daily = wb.create_sheet("Daily Activity Summary")
    ws_daily.sheet_properties.tabColor = _MED_BLUE

    # Build pivot: date → activity → {groups, delegates}
    pivot: dict[date, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {"groups": 0, "delegates": 0}))

    for _, row in df.iterrows():
        for col in act_cols:
            v = _col_val(row, col)
            if v is None:
                continue
            d = v if isinstance(v, date) else pd.Timestamp(v).date()
            act = col[4:]  # strip req_/opt_
            pivot[d][act]["groups"] += 1
            pivot[d][act]["delegates"] += int(row["GroupSize"])

    if smpw_df is not None and not smpw_df.empty:
        for _, row in smpw_df.iterrows():
            d = row["SMPW_Date"]
            if isinstance(d, str):
                d = date.fromisoformat(d)
            pivot[d]["SMPW"]["groups"] += 1
            pivot[d]["SMPW"]["delegates"] += int(row["GroupSize"])

    all_act_names = req_names + opt_names + (["SMPW"] if smpw_df is not None and not smpw_df.empty else [])

    # Write header row (date | act1 groups | act1 delegates | act2 groups | ...)
    header = ["Date"]
    for an in all_act_names:
        header += [f"{an} — Groups", f"{an} — Delegates"]
    header += ["Total Groups", "Total Delegates"]
    ws_daily.append(header)
    _style_header_row(ws_daily, 1, _MED_BLUE)

    for i, d in enumerate(all_dates):
        row_data = [str(d)]
        day_groups = 0
        day_delegates = 0
        for an in all_act_names:
            g = pivot[d][an]["groups"]
            dl = pivot[d][an]["delegates"]
            row_data += [g or "", dl or ""]
            day_groups += g
            day_delegates += dl
        row_data += [day_groups, day_delegates]
        ws_daily.append(row_data)
        for cell in ws_daily[ws_daily.max_row]:
            cell.font = _font(color="000000")
            if i % 2 == 0:
                cell.fill = _fill(_NEAR_WHITE)

    # Bold totals row
    ws_daily.append(["TOTAL"] + [""] * (len(header) - 1))
    _auto_col_widths(ws_daily)
    _freeze(ws_daily, "B2")

    # ── Sheet: Daily Activity by Hotel ─────────────────────────────────────────
    ws_hotel_daily = wb.create_sheet("Daily by Hotel")
    ws_hotel_daily.sheet_properties.tabColor = _LIGHT_BLUE

    # Build: date → hotel → activity → {groups, delegates}
    hotel_pivot: dict[date, dict[str, dict[str, dict]]] = defaultdict(
        lambda: defaultdict(lambda: defaultdict(lambda: {"groups": 0, "delegates": 0}))
    )

    for _, row in df.iterrows():
        hotel = str(row.get("Hotel") or "Unknown")
        for col in act_cols:
            v = _col_val(row, col)
            if v is None:
                continue
            d = v if isinstance(v, date) else pd.Timestamp(v).date()
            act = col[4:]
            hotel_pivot[d][hotel][act]["groups"] += 1
            hotel_pivot[d][hotel][act]["delegates"] += int(row["GroupSize"])

    if smpw_df is not None and not smpw_df.empty:
        for _, row in smpw_df.iterrows():
            d = row["SMPW_Date"]
            if isinstance(d, str):
                d = date.fromisoformat(d)
            hotel = str(row.get("Hotel") or "Unknown")
            hotel_pivot[d][hotel]["SMPW"]["groups"] += 1
            hotel_pivot[d][hotel]["SMPW"]["delegates"] += int(row["GroupSize"])

    hd_headers = ["Date", "Hotel"] + [f"{an} Grps" for an in all_act_names] + \
                 [f"{an} Del." for an in all_act_names] + ["Total Groups", "Total Delegates"]
    ws_hotel_daily.append(hd_headers)
    _style_header_row(ws_hotel_daily, 1, _LIGHT_BLUE)
    # fix font colour for light header
    for cell in ws_hotel_daily[1]:
        cell.font = _font(bold=True, color=_DARK_TEAL)

    hotel_names_sorted = sorted({str(row.get("Hotel") or "Unknown") for _, row in df.iterrows()})
    i = 0
    for d in all_dates:
        for hotel in hotel_names_sorted:
            if not any(hotel_pivot[d][hotel][an]["groups"] for an in all_act_names):
                continue
            row_data = [str(d), hotel]
            for an in all_act_names:
                row_data.append(hotel_pivot[d][hotel][an]["groups"] or "")
            for an in all_act_names:
                row_data.append(hotel_pivot[d][hotel][an]["delegates"] or "")
            tg = sum(hotel_pivot[d][hotel][an]["groups"] for an in all_act_names)
            td = sum(hotel_pivot[d][hotel][an]["delegates"] for an in all_act_names)
            row_data += [tg, td]
            ws_hotel_daily.append(row_data)
            for cell in ws_hotel_daily[ws_hotel_daily.max_row]:
                cell.font = _font(color="000000")
                if i % 2 == 0:
                    cell.fill = _fill(_NEAR_WHITE)
            i += 1

    _auto_col_widths(ws_hotel_daily)
    _freeze(ws_hotel_daily, "C2")

    # ── Sheet: Hotel Scheduling Coverage ──────────────────────────────────────
    ws_cov = wb.create_sheet("Hotel Coverage")
    ws_cov.sheet_properties.tabColor = _SAGE

    cov_headers = [
        "Hotel", "Total Groups", "Total Delegates",
        "Groups with Scheduling Days", "Delegates with Scheduling Days",
        "Groups with NO Scheduling",
    ]
    for an in req_names:
        cov_headers += [f"{an} — Scheduled", f"{an} — Rate"]
    if smpw_df is not None and not smpw_df.empty:
        cov_headers += ["SMPW Groups", "SMPW Delegates"]
    cov_headers += ["Avg Activities per Group"]

    ws_cov.append(cov_headers)
    _style_header_row(ws_cov, 1, _SAGE)
    for cell in ws_cov[1]:
        cell.font = _font(bold=True, color=_WHITE)

    for i, hotel in enumerate(hotel_names_sorted):
        hdf = df[df["Hotel"] == hotel]
        total_groups = len(hdf)
        total_del = int(hdf["GroupSize"].sum())

        # Has scheduling availability: group has at least one req_ scheduled
        has_any = hdf[[c for c in act_cols if c in hdf.columns and c.startswith("req_")]].notna().any(axis=1)
        sched_groups = int(has_any.sum())
        sched_del = int(hdf.loc[has_any, "GroupSize"].sum())
        no_sched = total_groups - sched_groups

        row_data = [hotel, total_groups, total_del, sched_groups, sched_del, no_sched]

        for an in req_names:
            col = f"req_{an}"
            n = int(hdf[col].notna().sum()) if col in hdf.columns else 0
            rate = f"{100*n//total_groups}%" if total_groups else "—"
            row_data += [n, rate]

        if smpw_df is not None and not smpw_df.empty:
            hsmpw = smpw_df[smpw_df["Hotel"] == hotel]
            row_data += [len(hsmpw), int(hsmpw["GroupSize"].sum())]

        # Average activities per group (count of req_ columns assigned)
        req_assigned = sum(
            int(hdf[c].notna().sum()) for c in act_cols if c in hdf.columns and c.startswith("req_")
        )
        avg = f"{req_assigned/total_groups:.2f}" if total_groups else "—"
        row_data.append(avg)

        ws_cov.append(row_data)
        for cell in ws_cov[ws_cov.max_row]:
            cell.font = _font(color="000000")
            if i % 2 == 0:
                cell.fill = _fill(_NEAR_WHITE)

        # Colour the "no scheduling" cell red if > 0
        no_sched_cell = ws_cov.cell(ws_cov.max_row, 6)
        if no_sched > 0:
            no_sched_cell.fill = _fill("FFE0DB")
            no_sched_cell.font = _font(bold=True, color=_RUST)

    _auto_col_widths(ws_cov)
    _freeze(ws_cov, "B2")

    # ── Sheet: Delegate Groups ─────────────────────────────────────────────────
    ws_groups = wb.create_sheet("Delegate Groups")
    group_cols = list(df.columns)
    _write_sheet(
        ws_groups, group_cols,
        [[_serialize(row[c]) for c in group_cols] for _, row in df.iterrows()],
        fill_hex=_DARK_TEAL,
    )

    # ── Sheet: Bus Assignments ─────────────────────────────────────────────────
    ws_buses = wb.create_sheet("Bus Assignments")
    if not bus_trips_df.empty:
        bus_cols = list(bus_trips_df.columns)
        _write_sheet(ws_buses, bus_cols,
                     [[_serialize(v) for v in row] for row in bus_trips_df.itertuples(index=False)])
    else:
        _write_sheet(ws_buses, ["No bus trips generated"], [])

    # Colour under-minimum rows
    if not bus_trips_df.empty:
        under_col_idx = bus_cols.index("UnderMinimum") + 1 if "UnderMinimum" in bus_cols else None
        if under_col_idx:
            for r in range(2, ws_buses.max_row + 1):
                if ws_buses.cell(r, under_col_idx).value is True:
                    for c in range(1, len(bus_cols) + 1):
                        ws_buses.cell(r, c).fill = _fill("FFE0DB")
                        ws_buses.cell(r, c).font = _font(color=_RUST)

    # ── Sheet: Non-Attendance ─────────────────────────────────────────────────
    ws_na = wb.create_sheet("Non-Attendance")
    if not non_attendance.empty:
        _write_sheet(ws_na, list(non_attendance.columns),
                     [[_serialize(v) for v in row] for row in non_attendance.itertuples(index=False)])
    else:
        _write_sheet(ws_na, ["Activity", "Note"], [
            [name, "All groups scheduled"] for name in required_activities
        ])

    # ── Sheet: SMPW Assignments ────────────────────────────────────────────────
    if smpw_df is not None and not smpw_df.empty:
        ws_smpw = wb.create_sheet("SMPW Assignments")
        ws_smpw.sheet_properties.tabColor = _AMBER
        _write_sheet(
            ws_smpw,
            ["GroupID", "Hotel", "Stop", "Delegates", "SMPW Date", "Fulfills FS"],
            [
                [row["GroupID"], row["Hotel"], row.get("TransportStop", ""),
                 row["GroupSize"], str(row["SMPW_Date"]), row["Fulfills_FS"]]
                for _, row in smpw_df.iterrows()
            ],
            fill_hex=_AMBER,
        )
        # fix header font
        for cell in ws_smpw[1]:
            cell.font = _font(bold=True, color=_DARK_TEAL)

    # ── Sheet: Activity Definitions ────────────────────────────────────────────
    ws_acts = wb.create_sheet("Activity Definitions")
    act_rows = []
    for name, defn in required_activities.items():
        cap = defn.get("capacity")
        act_rows.append(["Required", name, defn["priority"],
                         ", ".join(str(d) for d in defn["dates"]),
                         "Unlimited" if cap is None else str(cap), ""])
    for name, defn in optional_activities.items():
        cap = defn.get("capacity")
        act_rows.append(["Optional", name, "",
                         ", ".join(str(d) for d in defn["dates"]),
                         "Unlimited" if cap is None else str(cap),
                         f"{defn['want_pct']*100:.0f}%"])
    _write_sheet(ws_acts, ["Type", "Activity", "Priority", "Dates", "Capacity", "Want%"], act_rows)

    # ── Sheet: Assumptions ────────────────────────────────────────────────────
    ws_assum = wb.create_sheet("Assumptions")
    _write_sheet(ws_assum, ["#", "Assumption / Finding"],
                 [[i + 1, a] for i, a in enumerate(assumptions)])
    ws_assum.column_dimensions["B"].width = 100

    # ── Reorder sheets ─────────────────────────────────────────────────────────
    desired_order = [
        "Summary", "Daily Activity Summary", "Daily by Hotel",
        "Hotel Coverage", "Delegate Groups", "Bus Assignments",
        "Non-Attendance", "SMPW Assignments", "Activity Definitions", "Assumptions",
    ]
    for i, name in enumerate(desired_order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    wb.save(output_path)
    return output_path


def _serialize(v: Any) -> Any:
    if isinstance(v, date):
        return str(v)
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ""
    return v
